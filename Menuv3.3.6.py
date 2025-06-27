import os
import tkinter as tk
from tkinter import filedialog, messagebox
from PyPDF2 import PdfReader, PdfWriter
import pandas as pd
import aspose.pdf as ap
import re
import shutil
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import win32com.client as win32
from datetime import datetime
import openpyxl

class PDFApp(tk.Tk):

    
    def __init__(self):
        super().__init__()
        self.title("Convertidor de PDF a Excel")
        self.geometry("400x300")

        self.pdf_dir = tk.StringVar()
        self.generated_excel_files = []  # Lista para almacenar los archivos Excel generados
        self.generated_file_path = None
        self.sheets_with_data_bool = {} 
        

        self.btn_select_pdf = tk.Button(self, text="Seleccionar Directorio PDF", command=self.select_pdf_directory)
        self.btn_select_pdf.pack(pady=10)

        self.btn_process_files = tk.Button(self, text="Procesar PDFs", command=self.process_files, bg='green', fg='white')
        self.btn_process_files.pack(pady=20)
        self.btn_process_files.pack_forget()  # Initially hide this button

        self.btn_convert_to_excel = tk.Button(self, text="Convertir PDFs a Excel", command=self.convert_pages_to_excel, bg='blue', fg='white')
        self.btn_convert_to_excel.pack(pady=10)
        self.btn_convert_to_excel.pack_forget()  # Initially hide this button

        self.btn_unify_and_filter = tk.Button(self, text="Procesar Excels", command=self.unify_and_filter_excel, bg='orange', fg='white')
        self.btn_unify_and_filter.pack(pady=10)
        self.btn_unify_and_filter.pack_forget()

       

        self.btn_filter_excel = tk.Button(self, text="Filtrar Excels", command=self.filter_excel_files, bg='red', fg='white')
        self.btn_filter_excel.pack(pady=10)
        self.btn_filter_excel.pack_forget()

        self.btn_migrate_to_sheets = tk.Button(self, text="Pasar datos a Conversor de planillas", command=self.migrate_to_sheets, bg='purple', fg='white')
        self.btn_migrate_to_sheets.pack(pady=10)
        self.btn_migrate_to_sheets.pack_forget()

        self.btn_macros_to_sheets = tk.Button(self, text="Aplicar Macros", command=self.aplicar_macros, bg='green', fg='white')
        self.btn_macros_to_sheets.pack(pady=10)
        self.btn_macros_to_sheets.pack_forget()
        

    def select_pdf_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.pdf_dir.set(directory)
            self.btn_process_files.pack()  # Show the process button

    def process_files(self):
        if not self.pdf_dir.get():
            messagebox.showerror("Error", "Por favor, seleccione un directorio antes de procesar.")
            return
        
        pdf_directory = self.pdf_dir.get()
        self.pdf_pages_directory = os.path.join(pdf_directory, "pdfPages")
        
        
        os.makedirs(self.pdf_pages_directory, exist_ok=True)
        self.move_non_pdf_files(pdf_directory, self.pdf_pages_directory )
        self.split_all_pdfs(pdf_directory, self.pdf_pages_directory)

        messagebox.showinfo("Éxito", "Todos los PDFs han sido procesados exitosamente!")
        self.btn_process_files.pack_forget()  # Hide the process button
        self.btn_convert_to_excel.pack()  # Show the convert to Excel button

    
    
                

            
    def migrate_to_sheets(self):
            if self.generated_excel_files_directory:
                combined_dir = os.path.join(self.generated_excel_files_directory, "excelCombined")
                filtered_dir = os.path.join(combined_dir, "excelFiltered")  # Asegurarse que los archivos están en esta carpeta
                
                if os.path.exists(filtered_dir) and any(f.endswith('.xlsx') for f in os.listdir(filtered_dir)):
                    master_file_path = filedialog.askopenfilename(title="Seleccionar Archivo Maestro", defaultextension=".xlsm", filetypes=[("Excel Files", "*.xlsm")])
                    if not master_file_path:
                        messagebox.showerror("Error", "Por favor, seleccione el archivo maestro.")
                        return

                    save_file_path = filedialog.asksaveasfilename(title="Guardar Archivo Maestro", defaultextension=".xlsm", filetypes=[("Excel Files", "*.xlsm")])
                    if not save_file_path:
                        messagebox.showerror("Error", "Por favor, seleccione una ubicación y nombre para el archivo maestro.")
                        return

                    

                    self.update_master_sheet(master_file_path, filtered_dir, save_file_path)
                    self.generated_file_path = save_file_path
                    
                    self.run_macro(self.generated_file_path, "SumarCuotasVisa")
                    self.run_macro(self.generated_file_path, "SumarCuotasMastercard")
                    messagebox.showinfo("Éxito", "Datos enviados desde Excel a Conversor de planillas realizada con éxito.")
                    self.btn_migrate_to_sheets.pack_forget()
                    self.btn_macros_to_sheets.pack()
                    excel = win32.Dispatch("Excel.Application")
                    workbook=excel.Workbooks.Open(save_file_path)
                else:
                    messagebox.showerror("Error", "No se encontraron archivos Excel.")
            else:
                messagebox.showerror("Error", "No se ha establecido un directorio de archivos Excel generados.")

    def aplicar_macros(self):
        if self.generated_file_path:
            for sheet, has_data in self.sheets_with_data_bool.items():
                if has_data:
                    macro_name = f"{sheet.replace(' ', '_')}"
                    self.run_macro(self.generated_file_path, macro_name)
                    print(f"Macro {macro_name} ejecutada para {sheet}")
            self.run_macro(self.generated_file_path, "OpAnticipo")
            self.run_macro(self.generated_file_path, "AplicarFormulas")
            self.run_macro(self.generated_file_path, "OperacionesAgregar")
            self.run_macro(self.generated_file_path, "NroAutorizacion")
            self.run_macro(self.generated_file_path, "CopiarDatosEspeciales")
            print(f"Macros finales aplicadas")
            self.btn_macros_to_sheets.pack_forget()
        else:
            messagebox.showerror("Error", "No se ha generado ningún archivo para aplicar macros.")

    def adapt_formula(self, formula, new_row):
        cell_ref_pattern = r'(\$?[A-Za-z]+)(\$?\d+)'

        def replace_cell_reference(match):
            col_letter, row_number = match.group(1), match.group(2)
            if '$' in col_letter or '$' in row_number:
                return f"{col_letter}{row_number}"
            else:
                adjusted_row_number = new_row if int(row_number.lstrip('$')) == 2 else int(row_number) + new_row - 2
                return f"{col_letter}{adjusted_row_number}"

        formula = re.sub(r'([A-Z]+)(\d+)', lambda x: replace_cell_reference(x), formula)
        return formula

    def run_macro(self, excel_path, macro_name):
        excel = win32.DispatchEx("Excel.Application")
        workbook = excel.Workbooks.Open(excel_path)
    
    # Mantener la visibilidad de la aplicación independiente
        excel.Visible = True

    # Ejecutar la macro
        excel.Application.Run(f"{macro_name}")
    
    # Guardar y cerrar el archivo
        workbook.Save()
        workbook.Close()
    
    # Cerrar la instancia de Excel después de terminar
        excel.Quit()

    



    def move_non_pdf_files(self, src_directory, dest_directory):
        non_pdf_extensions = ['.csv', '.xls', '.xlsx']
        for item in os.listdir(src_directory):
            if any(item.endswith(ext) for ext in non_pdf_extensions):
                src_path = os.path.join(src_directory, item)
                dest_path = os.path.join(dest_directory, item)
                shutil.move(src_path, dest_path)  # Mover el archivo
                print(f"Archivo {item} movido a {dest_directory}")

    def update_master_sheet(self, master_path, folder_path, save_path):
        if not os.path.exists(save_path):
            
            shutil.copy(master_path, save_path)
            print("Copia creada en:", save_path)
            
            self.run_macro(save_path, "Limpiar")
            self.run_macro(save_path, "LimpiarAMEX2")
            self.run_macro(save_path, "LimpiarAmex")
            self.run_macro(save_path, "LimpiarMastercard")
            self.run_macro(save_path, "LimpiarMastercard_Debito")
            self.run_macro(save_path, "LimpiarVisa")
            self.run_macro(save_path, "LimpiarVisa_Debito")
            self.run_macro(save_path, "LimpiarMaestro")
            self.run_macro(save_path, "LimpiarQR")
            self.run_macro(save_path, "LimpiarCabal")
            self.run_macro(save_path, "LimpiarArgencard")
        else:
            print("Copia ya existe en:", save_path)
        wb_maestro = openpyxl.load_workbook(save_path, keep_vba=True)



        name_to_sheet = {
            'Visa Débito': 'Visa debito',
            'Visa Crédito': 'Visa',
            'Mastercard Debit': 'Mastercard debito',
            'Maestro': 'MAESTRO',
            'Mastercard Crédito': 'Mastercard',
            'Resumen': 'CABAL',
            'Amex': 'AMEX FISERV',
            'Argencard': 'ARGENCARD',
            'Detalles': 'AMEX_2',
            'Pagos': 'QR',
            'NaranjaX':'Naranja'
        }
        

        sheets_with_data = {}
         

        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            if os.path.isfile(file_path):
                for key, value in name_to_sheet.items():
                    if key in filename:
                        sheet_name = value
                        break
                else:
                    continue

                wb_current = openpyxl.load_workbook(file_path, data_only=False)
                sheet_current = wb_current.active
                sheet_maestro = wb_maestro[sheet_name]

                if sheet_maestro.max_row > 2:
                    sheet_maestro.delete_rows(3, sheet_maestro.max_row - 1)

                formulas = {}
                for col_idx in range(1, 47):  
                    cell = sheet_maestro.cell(row=2, column=col_idx)
                    if cell.data_type == 'f':
                        formulas[col_idx] = cell.value

                data_added = False
                for row_idx, row in enumerate(sheet_current.iter_rows(min_row=2, max_col=47), start=2):
                    for col_idx, cell in enumerate(row, start=1):
                        new_cell = sheet_maestro.cell(row=row_idx, column=col_idx)
                        if col_idx in formulas:
                            adapted_formula = self.adapt_formula(formulas[col_idx], row_idx)
                            new_cell.value = '=' + adapted_formula
                        else:
                            new_cell.value = cell.value
                        if cell.value:
                            data_added = True

                        if row_idx > 2 and col_idx > 2:
                            prev_cell = sheet_maestro.cell(row=row_idx - 1, column=col_idx)
                            new_cell.number_format = prev_cell.number_format
                            new_cell.font = copy(prev_cell.font)
                            new_cell.alignment = copy(prev_cell.alignment)
                            new_cell.border = copy(prev_cell.border)
                            new_cell.fill = copy(prev_cell.fill)

                sheets_with_data[sheet_name] = data_added
                self.sheets_with_data_bool[sheet_name] = data_added  
                
                last_row_with_data = sheet_maestro.max_row
                table_name = self.get_table_name_for_sheet(sheet_name)
                if table_name:
                    table = sheet_maestro.tables.get(table_name)
                    if table:
                        table_range = table.ref.split(':')
                        table_ref_start = table_range[0]
                        table_ref_end = table_range[1]
                        updated_table_range = f"{table_ref_start}:{table_ref_end.split('$')[0]}{last_row_with_data}"
                        if re.match(r'^[$]?([A-Za-z]{1,3})[$]?(\d+)(:[$]?([A-Za-z]{1,3})[$]?(\d+)?)?$|^[A-Za-z]{1,3}:[A-Za-z]{1,3}$', updated_table_range):
                            table.ref = updated_table_range
                        else:
                            print(f"La referencia de la tabla '{table_name}' no es válida: '{updated_table_range}'")


        wb_maestro.save(save_path)
        
        
        
        

        return sheets_with_data, self.sheets_with_data_bool



    
    def get_table_name_for_sheet(self, sheet_name):
        table_names = {
            'Visa debito': 'Tabla14',
            'Visa': 'Tabla1',
            'Mastercard debito': 'Tabla145',
            'MAESTRO': 'Tabla1456',
            'Mastercard': 'Tabla13',
            'CABAL': 'Tabla7',
            'AMEX FISERV': 'Tabla19',
            'ARGENCARD': 'Tabla137',
            'AMEX_2': 'Tabla911',
            'QR':'Tabla1412',
            'Naranja':'Tabla113'
        }
        return table_names.get(sheet_name, None)

    def auto_adjust_column_width(self, sheet):
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                         max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    def unify_and_filter_excel(self):
        if os.path.exists(self.generated_excel_files_directory) and any(os.path.isfile(os.path.join(self.generated_excel_files_directory, f)) for f in os.listdir(self.generated_excel_files_directory)):
            combined_dir = os.path.join(self.generated_excel_files_directory, "excelCombined")
            os.makedirs(combined_dir, exist_ok=True)
            self.convert_xls_to_xlsx(self.generated_excel_files_directory)
            
            self.combine_and_save_by_tarjeta(self.generated_excel_files_directory, combined_dir)
            
            

            messagebox.showinfo("Éxito", "Archivos Excel unificados exitosamente!")
            self.move_non_pdf_files(self.generated_excel_files_directory, combined_dir)
            self.btn_unify_and_filter.pack_forget()
            self.btn_filter_excel.pack()
        else:
            messagebox.showerror("Error", "No se han generado archivos Excel para unificar y filtrar.")

    def combine_and_save_by_tarjeta(self, directory, output_dir):
        nombres_tarjetas = ["VisaDébito", "VisaCrédito", "MastercardDebit", "MastercardCrédito", "NaranjaX", "Argencard", "Detalles", "Amex", "Maestro", "Resumen", "Pagos"]
        dfs_tarjetas = {nombre: pd.DataFrame() for nombre in nombres_tarjetas}
        self.process_resumen_files(directory, dfs_tarjetas, output_dir)
        self.process_special_cases(directory, dfs_tarjetas, output_dir)
        
        # Mensajes de depuración
        print("Archivos en el directorio inicial:")
        for file in os.listdir(directory):
            print(f"Archivo: {file}")
        
        # Procesar los archivos
        for file in os.listdir(directory):
            file_path = os.path.join(directory, file)
            if file.endswith(('.xlsx', '.xls', '.csv')) and 'Resumen' not in file_path and 'Detalles' not in file_path:
                for nombre_tarjeta in nombres_tarjetas:
                    if self.nombre_en_archivo(nombre_tarjeta, file):
                        print(f"Procesando archivo para {nombre_tarjeta}: {file_path}")
                        df_temp = pd.read_excel(file_path, header=None)
                        df_temp = df_temp.dropna(how='all')
                        print(f"Filas en {file}: {len(df_temp)}")
                        
                        # Agregar los datos al DataFrame de la tarjeta correspondiente
                        dfs_tarjetas[nombre_tarjeta] = pd.concat([dfs_tarjetas[nombre_tarjeta], df_temp], ignore_index=True)
                    else:
                        print(f"Archivo {file} no coincide con {nombre_tarjeta}.")
        
        # Guardar los DataFrames combinados
        for nombre_tarjeta, df in dfs_tarjetas.items():
            print(f"DataFrame de {nombre_tarjeta}: {len(df)} filas.")
            if not df.empty:
                output_path = os.path.join(output_dir, f"{nombre_tarjeta}_combined.xlsx")
                df.to_excel(output_path, index=False)
                print(f"Archivo guardado: {output_path}")
            else:
                print(f"No se encontraron datos para {nombre_tarjeta}.")

        # Opcional: eliminar archivos originales solo después de haber guardado los combinados
        print("Eliminando archivos procesados...")
        self.cleanup_files(directory)

    
    def process_resumen_files(self, directory, dfs_tarjetas, output_dir):
        for file in os.listdir(directory):
            file_path = os.path.join(directory, file)
            if file.endswith(('.xlsx', '.xls', '.csv')) and 'Resumen' in file_path:
                print(f"Procesando archivo de Resumen: {file_path}")
                df_temp = pd.read_excel(file_path, header=None)
                df_temp = df_temp.dropna(how='all')
                dfs_tarjetas["Resumen"] = pd.concat([dfs_tarjetas["Resumen"], df_temp])

        # Guardar el DataFrame de Resumen combinado
        if not dfs_tarjetas["Resumen"].empty:
            output_path = os.path.join(output_dir, "Resumen_combined.xlsx")
            dfs_tarjetas["Resumen"].to_excel(output_path, index=False)
            print(f'Archivo de Resumen guardado: {output_path}')

    def process_special_cases(self, directory, dfs_tarjetas, output_dir):
        for file in os.listdir(directory):
            file_path = os.path.join(directory, file)
            if file.endswith(('.xlsx', '.xls', '.csv')) and 'Detalles' in file_path:
                df_header = pd.read_excel(file_path, header=None, nrows=11)

                if df_header.shape[0] > 10 and df_header.shape[1] > 21:
                    value_to_copy = df_header.iat[10, 21]
                    df_temp = pd.read_excel(file_path, header=None, skiprows=14)
                    df_temp = df_temp.dropna(how='all')

                    if df_temp.shape[1] > 17:
                        last_row = df_temp.iloc[:, 0].last_valid_index()  # Última fila válida basada en la columna A (índice 0)
                        if last_row is not None and last_row + 14 >= 15:  # Ajustar por skiprows
                            df_temp.loc[15 - 14:last_row, 17] = value_to_copy  # Ajustar índice para skiprows
                else:
                    df_temp = pd.read_excel(file_path, header=None, skiprows=14)
                    df_temp = df_temp.dropna(how='all')
                rows = df_temp.values.tolist()
                for row in rows:
                    while row[0]=="":
                        row=row[1:]+[""]
                          
                        if not any(row):
                            break

                dfs_tarjetas["Detalles"] = pd.concat([dfs_tarjetas["Detalles"], df_temp])
        if not dfs_tarjetas["Detalles"].empty:
            output_path = os.path.join(output_dir, "Detalles_combined.xlsx")
            dfs_tarjetas["Detalles"].to_excel(output_path, index=False)
            print(f'Archivo de Detalles guardado: {output_path}')

    def cleanup_files(self, directory):
        for file in os.listdir(directory):
            file_path = os.path.join(directory, file)
            if file.endswith(('.xlsx', '.xls', '.csv')):
                try:
                    os.remove(file_path)
                    print(f'Archivo {file_path} eliminado correctamente.')
                except Exception as e:
                    print(f'No se pudo eliminar {file_path}: {e}')

    def nombre_en_archivo(self, nombre_tarjeta, nombre_archivo):
       # Crear un patrón para buscar la tarjeta en el nombre del archivo
        pattern = rf"{re.escape(nombre_tarjeta)}"
        # Imprimir el patrón para depuración
        print(f"Buscando patrón: '{pattern}' en archivo: '{nombre_archivo}'")
        return re.search(pattern, nombre_archivo) is not None

    def split_all_pdfs(self, pdf_directory, pdf_pages_directory):
        for filename in os.listdir(pdf_directory):
            if filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(pdf_directory, filename)
                self.split_pdf_pages(pdf_path, pdf_pages_directory)

    def split_pdf_pages(self, pdf_path, output_directory):
        with open(pdf_path, 'rb') as file:
            pdf_reader = PdfReader(file)
            num_pages = len(pdf_reader.pages)
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                output_page_path = os.path.join(output_directory, f"{os.path.splitext(os.path.basename(pdf_path))[0]}_página_{page_num + 1}.pdf")
                pdf_writer = PdfWriter()
                pdf_writer.add_page(page)
                with open(output_page_path, 'wb') as output_file:
                    pdf_writer.write(output_file)

    def convert_pages_to_excel(self):

        excel_directory = os.path.join(os.path.dirname(self.pdf_pages_directory), "pagesExcel")
        os.makedirs(excel_directory, exist_ok=True)
        self.move_non_pdf_files(self.pdf_pages_directory, excel_directory)

        self.process_pdfs_from_directory(self.pdf_pages_directory, excel_directory)
        
        messagebox.showinfo("Éxito", "Todas las páginas han sido convertidas a Excel exitosamente!")
        self.btn_convert_to_excel.pack_forget()
        self.generated_excel_files_directory = excel_directory
        self.btn_unify_and_filter.pack()

    def process_pdfs_from_directory(self, pdf_directory, excel_directory):
        for filename in os.listdir(pdf_directory):
            if filename.lower().endswith('.pdf'):
                pdf_path = os.path.join(pdf_directory, filename)
                excel_filename = filename[:-4] + '.xlsx'
                excel_path = os.path.join(excel_directory, excel_filename)
                self.convert_pdf_to_excel(pdf_path, excel_path)

    

    def convert_pdf_to_excel(self, pdf_path, excel_path):
        try:
            document = ap.Document(pdf_path)
            excel_save_options = ap.ExcelSaveOptions()
            excel_save_options.format = ap.ExcelSaveOptions.ExcelFormat.XLSX
            document.save(excel_path, excel_save_options)
            print(f'{pdf_path} convertido a Excel con Aspose en {excel_path}')

            if "Maestro" in pdf_path.upper():
                self.adjust_columns_excel(excel_path)

            os.remove(pdf_path)
            print(f'Archivo PDF original {pdf_path} eliminado exitosamente.')

        except Exception as e:
            print(f"Error al convertir {pdf_path} a Excel: {e}")

    def adjust_columns_excel(self, excel_path):
        try:
            data = pd.read_excel(excel_path, header=0)
            if data.shape[1] >= 8:
                data.iloc[:, 2] =data.iloc[:,2].astype(str) + ' ' + data.iloc[:, 3].astype(str) + ' ' + data.iloc[:, 4].astype(str)
                data.iloc[:, 3] = data.iloc[:, 5]
                data.iloc[:, 4] = data.iloc[:, 6]
                data.iloc[:, 5] = data.iloc[:, 7]
                data.drop(data.columns[[6, 7]], axis=1, inplace=True)
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    data.to_excel(writer, index=False)
                print(f'Ajuste de columnas realizado en {excel_path}')
        except Exception as e:
            print(f"Error al ajustar las columnas en {excel_path}: {e}")
            
 

    def convert_to_float(self,x):
        if pd.notna(x) and isinstance(x, str):
            try:
                return float(x.replace('.', '').replace(',', '.'))
            except ValueError:
                return x  
        return x

    def separate_lote_cupon(self, data):
        split_content = data.iloc[:, 2].str.split(expand=True)
        data.insert(3, 'lote', split_content[1])
        data.insert(4, 'cupon', split_content[2]) if split_content.shape[1] > 1 else None
        data.iloc[:, 2] = split_content[0] if split_content.shape[1] > 2 else None
        return data

    def clean_empty_cells(self, data, start_col, sheet_name):
        for col in range(start_col, len(data.columns)):
            data.iloc[:, col] = data.iloc[:, col].apply(lambda x: x if pd.notna(x) and str(x).strip() != '' else None)
        if 'Maestro' in sheet_name:
            data['Term/Lote/Cupon'] = data['Term/Lote/Cupon'].apply(lambda x: x if pd.notna(x) and str(x).strip() != '' else "Default Value")
        return data

    def apply_numeric_conversion(self, data):
        for col in data.columns:
            data[col] = data[col].apply(self.convert_to_float)
        return data
    def apply_dateConverter(self, data, date_column_name, date_format="%d/%m/%Y"):
        
        try:
            data[date_column_name] = pd.to_datetime(data[date_column_name], format=date_format, errors='coerce').dt.date
            print(f"Conversion successful for column: {date_column_name}")
        except Exception as e:
            print(f"Error during date conversion in column {date_column_name}: {str(e)}")
        return data
        
    def adjust_card_number(self, data):
        data['Número de Tarjeta'] = data['Número de Tarjeta'].apply(lambda x: x[-4:] if pd.notna(x) and len(x) > 4 else x)
        return data
    def clean_plan_column(self, value):
        if pd.isna(value):
            return None
        # Eliminar signos de pesos y espacios
        cleaned_value = value.replace('$', '').strip()
        # Convertir "1 z" a "3"
        if cleaned_value == "1 z":
            return "3"
        return cleaned_value

    def filter_excel(self, file_path):

        fixed_headers = [
    "Trx", "Fecha Pres Fecha", "Term/Lote/Cupon", "Tarj", "Plan Cuota", "T F", 
    "T.N.A. %", "Ventas con/Dto.", "Ventas sin/Dto.", "Dto. Arancel", 
    "Dto. Financ.", "Cod. Rechazo Mot. contrap."
]
        headers_american=[
    "Fecha de la venta", "Número de la transacción", "Importe de descuento de la transacción", "Importe", "Número de referencia de cargo", "Número de Tarjeta",
    "Alias", "Número del Establecimiento", "Número de la factura del Sumario de Cargos", "Tipo", "Número de boleto de aerolínea", "Número de acuerdo de alquiler", "Identificador de referencia",
    "Motivo de rechazo", "Número de Terminal", "Número de referencia", "Fecha de liquidación", "Numero de cuotas"
    ]
        headers_qr=['Fecha Operación', 'Fecha de Presentación', 'Fecha de Pago', 'Nro. de Cupon',	'Nro.', 'de', 'Comercio', 'Moneda', 'Total Bruto', 'Total Descuento', 'Total Neto', 'Entidad Pagadora', 'Cuenta Bancaria', 'Nro. Liquidación', 'Nro. de Lote', 'Tipo de Liquidacion', 'Estado', 'Cuotas', 'Nro. de Autorizacion', 'Tarjeta', 'Tipo de Operacion', 'Comercio Participante', 'Promoción Plan']
        headers_cabal=['Producto', 'Lote', 'Terminal', 'Fec. Presentacion', 'Cantidad', 'Total Ventas', 'Arancel', 'Neto' ]

        
        try:
            data = pd.read_excel(file_path, header=0)
            if 'NaranjaX' in file_path:
                header_row = data.index[data.iloc[:, 0] == 'Fecha de Compra'].tolist()
                if header_row:
                    data.columns = data.iloc[header_row[0]]
                    data = data[header_row[0]+1:]  

                data = data[data['Operac.'].apply(lambda x: not (pd.isna(x) or str(x).strip() == ''))]
                if 'Terminal-Lote' in data.columns:
                    def split_terminal_lote(value):
                        if pd.isna(value):
                            return None, None
                        # Convertir a cadena en caso de que el valor no sea una cadena
                        value_str = str(value)
                        if '-' in value_str:
                            terminal, lote = value_str.split('-')
                            return terminal, lote
                        return value_str, None

                    # Aplicar la función y crear nuevas columnas
                    terminal_lote = data['Terminal-Lote'].apply(split_terminal_lote)
                    data.insert(2, 'Terminal', terminal_lote.apply(lambda x: x[0]))
                    data.insert(3, 'Lote', terminal_lote.apply(lambda x: x[1]))
                    
                    # Eliminar la columna original 'Terminal-Lote'
                    data.drop('Terminal-Lote', axis=1, inplace=True)
                if 'Plan' in data.columns:
                    data['Plan'] = data['Plan'].apply(self.clean_plan_column)
                
                filtered_data = data
                return filtered_data

            if 'Pagos' in file_path:
                data = pd.read_excel(file_path, header=0)
                data = self.apply_dateConverter(data, 'FECHA DE OPERACIÓN', "%d/%m/%Y")
                filtered_data = data
                return filtered_data
                
            if 'Detalles' in file_path:
                if data.shape[1] > 18:
                    data = data.iloc[:, :-4]
                data.columns = headers_american
                data = data[data.iloc[:, 0] != 'Fecha de la venta']
                data = self.adjust_card_number(data)
                data = self.apply_dateConverter(data, 'Fecha de la venta', "%d/%m/%Y")
                filtered_data=data
                return filtered_data
            if 'Resumen' in file_path:
                data.columns = headers_cabal
                data = data.astype(str)

                condition = (data.iloc[:, 0].str.contains("CABAL DEBITO|TARJETAS CABAL", case=False, regex=True) &
                         ~(data.iloc[:, 1].str.contains("Sub Total", na=False, case=False)))

                data = data[condition]
                data = self.apply_numeric_conversion(data)
                data = self.apply_dateConverter(data, 'Fec. Presentacion', "%d/%m/%Y")
                filtered_data=data
                return filtered_data
                
            else:
                if data.shape[1] > 12:
                    data = data.iloc[:, :-1]
                data.columns = fixed_headers[:data.shape[1]]
                #if 'Maestro' in file_path:
                    #data = self.separate_lote_cupon(data)
                    #data = self.clean_empty_cells(data, 2, 'MAESTRO')
               # else:
                data = self.clean_empty_cells(data, 3, 'Other')

                
                data = self.apply_numeric_conversion(data)
                data = self.apply_dateConverter(data, 'Fecha Pres Fecha', "%d/%m/%Y")
                if data.iloc[:, 0].isin(['Plan cuota', 'Venta ctdo', 'Contr ctdo', 'Devol ctdo', 'Contr cuo', 'Devol cuo', 'Repres cuo', 'Repres ctdo']).any():
                    #repres cuo, ctdo puede chotearse
                    filtered_data = data[data.iloc[:, 0].isin(['Plan cuota', 'Venta ctdo','Contr ctdo', 'Devol ctdo', 'Contr cuo', 'Devol cuo', 'Repres cuo', 'Repres ctdo'])]
                    return filtered_data
                else:
                    print(f"No se encontraron entradas válidas en la primera columna del archivo: {file_path}")
                    return None
        except Exception as e:
            print(f"Error procesando el archivo {file_path}: {e}")
            return None

    def filter_excel_files(self):
        combined_dir=os.path.join(self.generated_excel_files_directory, 'excelCombined')
        self.convert_xls_to_xlsx(combined_dir)
        if os.path.exists(combined_dir):
            output_folder_path = os.path.join(combined_dir, 'excelFiltered')
            os.makedirs(output_folder_path, exist_ok=True)
            self.filter_combined_excel_files(combined_dir, output_folder_path)
            messagebox.showinfo("Éxito", "Archivos Excel combinados filtrados exitosamente!")
            self.btn_filter_excel.pack_forget()
            self.btn_migrate_to_sheets.pack()
        else:
            messagebox.showerror("Error", "No se encontró la carpeta de archivos combinados.")
    def filter_combined_excel_files(self, combined_dir, output_folder):
        combined_files = [f for f in os.listdir(combined_dir) if f.endswith('.xlsx')]
        for combined_file in combined_files:
            combined_file_path = os.path.join(combined_dir, combined_file)
            filtered_data = self.filter_excel(combined_file_path)
            if filtered_data is not None and 'NaranjaX' in combined_file:  
                output_path = os.path.join(output_folder, f'filtered_{combined_file}')
                self.save_excel_with_left_aligned_headers(filtered_data, output_path)
                print(f'Archivo filtrado y formateado guardado en: {output_path}')
            elif filtered_data is not None:
                output_path = os.path.join(output_folder, f'filtered_{combined_file}')
                filtered_data.to_excel(output_path, index=False)
                print(f'Archivo combinado filtrado guardado en: {output_path}')

            os.remove(combined_file_path)
            print(f'Archivo original {combined_file_path} eliminado exitosamente.')


    def save_excel_with_left_aligned_headers(self, dataframe, file_path):
        dataframe.to_excel(file_path, index=False)
        workbook = load_workbook(file_path)
        worksheet = workbook.active


        for col in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal='left')


        workbook.save(file_path)

    
    def convert_xls_to_xlsx(self, folder_path):
        files = [f for f in os.listdir(folder_path) if f.endswith('.xls') or f.endswith('.csv')]
        for file in files:
            file_path = os.path.join(folder_path, file)
            new_file_path = os.path.join(folder_path, os.path.splitext(file)[0] + '.xlsx')
            try:
                if file.endswith('.xls'):
                    # Para archivos XLS
                    data = pd.read_excel(file_path)
                    data.to_excel(new_file_path, index=False, engine='openpyxl')
                    print(f"Converted XLS to XLSX: {new_file_path}")
                else:
                    data = pd.read_csv(file_path, delimiter=';')  
                    wb = Workbook()
                    ws = wb.active

                    
                    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)

                    
                    for col in ws.columns:
                        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                        col_letter = col[0].column_letter
                        ws.column_dimensions[col_letter].width = max_length + 2

                    wb.save(new_file_path)
                    print(f"Converted CSV to XLSX with formatting: {new_file_path}")
                os.remove(file_path)
                print(f"Archivo original {file_path} fue eliminado.")


            except Exception as e:
                print(f"Error converting {file_path}: {e}")



        
        
if __name__ == "__main__":
    app = PDFApp()
    app.mainloop()

